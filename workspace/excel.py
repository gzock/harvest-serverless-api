#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fonts import Font
#from openpyxl.drawing.image import Image
from PIL import Image

if __name__ == '__main__':
  wb = openpyxl.load_workbook('./sample.xlsx')
  ws = wb['photos']

  for row in ws:
    for cell in row:
        if cell.value == "photo":
          font = Font(
            name='Arial',
            sz=10,
          )
          cell.font = font
          # https://stackoverflow.com/questions/51135083/openpyxl-how-to-iterate-over-sheets-to-add-an-image-inside-each-one
          #img = Image('sample.jpg')
          print(cell.value)
          print(ws.row_dimensions[cell.row].height)
          print(ws.column_dimensions[get_column_letter(cell.column)].width)

          ratio = 1
          height_ratio = 1

          cellHeight = round(ws.row_dimensions[cell.row].height * 1.33)
          cellWidth = round(ws.column_dimensions[get_column_letter(cell.column)].width * 7)
          print("cell size:")
          print(cellHeight)
          print(cellWidth)

          img = Image.open('sample.jpg')
          print("original image size:")
          print(img.height)
          print(img.width)
          ratio = cellWidth / img.width
          width = img.width * ratio
          height = img.height * ratio
          print(ratio)
          print(height)
          print(width)

          if height > cellHeight:
            height_ratio = cellHeight / height
            print(height_ratio)
            height *= height_ratio
            width *= height_ratio

          print(width)
          print(height)

          img = img.resize((round(width), round(height)),Image.NEAREST)
          print("resized image:")
          print(img.height)
          print(img.width)
          img.save('resize.jpg')

          img = openpyxl.drawing.image.Image('resize.jpg')
          anchor = get_column_letter(cell.column) + str(cell.row)
          #print(anchor)
          #print(cell.font)
          #print(cell.font.charset)
          #print(cell.font.vertAlign)
          #print(cell.font)
          #print(dir(cell.font))
          ws.add_image(img, anchor)

  # 保存
  wb.save('pasted.xlsx')

#
#wb = openpyxl.Workbook()
#ws = wb.worksheets[0]
#img = openpyxl.drawing.image.Image('photo.jpg')
#ws.add_image(img,'F10')
