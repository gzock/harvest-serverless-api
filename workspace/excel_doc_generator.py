#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fonts import Font
from PIL import Image

class ExcelDocGenerator():

  def __init__(self, workbook, worksheet):
    if workbook:
      self.load_worksheet(workbook)
    if worksheet:
      self.active_worksheet(worksheet)

  def load_workbook(self, filename):
    self.wb = openpyxl.load_workbook(filename)

  def active_worksheet(self, sheet_name):
    if self.wb:
      self.ws = self.wb[sheet_name]

  def get_cell_height(self, cell):
    return round(self.ws.row_dimensions[cell.row].height)

  def get_cell_width(self, cell):
    return round(self.ws.column_dimensions[get_column_letter(cell.column)].width)

  def get_cell_height_px(self, cell):
    return round(self.get_cell_height(cell) * 1.33)

  def get_cell_width_px(self, cell):
    return round(self.get_cell_width(cell) * 7)

  def save(self, filename):
    return self.wb.save(filename)

  def open_image(self, filename):
    return  Image.open(filename)

  def save_image(self, img, filename):
    return img.save(filename)

  def resize_image(self, img, width, height):
    return img.resize((width, height),Image.NEAREST)

  def resize_image_to_cell(self, img, cell):
    ratio = 1
    width = img.width
    height = img.height

    font = Font(
      name='Arial',
      sz=10,
    )
    cell.font = font

    cellHeight = self.get_cell_height(cell)
    cellWidth = self.get_cell_width(cell)

    if width > cellWidth:
      ratio = cellWidth / width
      width *= ratio
      height *= ratio

    if height > cellHeight:
      ratio = cellHeight / height
      height *= ratio
      width *= ratio

    return self.resize_image(img, round(width), round(height))

  def add_image(self, filename, cell, adjust=True):
    if adjust:
      img = Image.open(filename)
      img = self.resize_image_to_cell(img, cell)
      self.save_image(img, '/tmp/resize.jpg')
      img = openpyxl.drawing.image.Image('/tmp/resize.jpg')
    else:
      img = openpyxl.drawing.image.Image(filename)

    anchor = get_column_letter(cell.column) + str(cell.row)
    return self.ws.add_image(img, anchor)

  def collect_cells(self, value):
    cells = []
    for row in self.ws:
      for cell in row:
        if cell.value == value:
          cells.append(cell)
    else:
      return cells

  def add_each_images(self, images, cells):
    for image in images:
      self.add_image(image, cells.pop(0))

  def generate(self, images, save_filename):
    cells = self.collect_cels("photo")
    self.add_each_images(images, cells)
    return self.save(save_filename)

